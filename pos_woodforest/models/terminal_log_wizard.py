import logging
import json
import base64
import uuid
from io import BytesIO
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from odoo import models, fields, api
from odoo.exceptions import UserError
from ..controllers.main import (
    build_url, build_header_hash, deep_clean_payload,
)
from ..services.logging_service import log_payrillium_event

_logger = logging.getLogger(__name__)


# ── Friendly error mapping ────────────────────────────────────────
_FRIENDLY_ERRORS = {
    "Read timed out": (
        "The terminal did not respond within the allowed time.\n\n"
        "Possible causes:\n"
        "  • The terminal is powered off or disconnected from the network.\n"
        "  • Weak or unstable internet connection on the terminal.\n\n"
        "Please verify the terminal is online (try 'Check Terminal' first) "
        "and then retry the log download."
    ),
    "Connection refused": (
        "Could not connect to the terminal service.\n\n"
        "The terminal may be offline or undergoing maintenance.\n"
        "Please try again in a few minutes."
    ),
    "Name or service not known": (
        "Could not resolve the terminal service address.\n\n"
        "This usually indicates a DNS or network issue on the server side.\n"
        "Please check your internet connection and try again."
    ),
    "Connection reset": (
        "The connection to the terminal was interrupted.\n\n"
        "Please try again. If the problem persists, check the terminal's "
        "network connection."
    ),
}


def _friendly_message(raw_error):
    """Return a user-friendly message for known error patterns."""
    raw = str(raw_error)
    for pattern, message in _FRIENDLY_ERRORS.items():
        if pattern in raw:
            return message
    return f"Unexpected error: {raw}"


class TerminalLogWizard(models.TransientModel):
    _name = "payrillium.terminal.log.wizard"
    _description = "Download Terminal Logs"

    terminal_id = fields.Many2one(
        "payrillium.terminal", string="Terminal",
        required=True, readonly=True,
    )
    date = fields.Date(
        string="Date", required=True,
        default=fields.Date.context_today,
    )

    # ── core ────────────────────────────────────────────────────────
    def action_download(self):
        self.ensure_one()
        terminal = self.terminal_id
        if not terminal.serial:
            raise UserError("Terminal has no serial number configured.")

        exec_id = f"tlog_{terminal.serial}_{uuid.uuid4().hex[:8]}"
        date_str = self.date.strftime("%Y-%m-%d")

        _logger.info(
            "Downloading logs for terminal %s (%s), exec_id=%s, date=%s",
            terminal.name, terminal.serial, exec_id, date_str,
        )

        # ── Fetch logs ────────────────────────────────────────────
        try:
            entries = self._fetch_logs_for_date(terminal, date_str, exec_id)
        except requests.exceptions.RequestException as e:
            friendly = _friendly_message(e)
            raise UserError(friendly)

        if not entries:
            return {
                "type": "ir.actions.client",
                "tag": "display_notification",
                "params": {
                    "title": "No Logs",
                    "message": (
                        f"No log entries found for {date_str} "
                        f"on terminal '{terminal.name}'."
                    ),
                    "type": "warning",
                    "sticky": False,
                },
            }

        # ── Build Excel workbook ──────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Terminal Logs"

        columns = [
            "#", "Date / Time", "Route", "Source",
            "Success", "Request", "Response",
        ]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid",
        )
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, entry in enumerate(entries, 2):
            epoch_ms = entry.get("date", 0)
            date_val = (
                datetime.utcfromtimestamp(epoch_ms / 1000).strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
                if epoch_ms
                else ""
            )

            route = (entry.get("route") or "").strip()
            source = entry.get("sourceType", "")
            response = entry.get("response", {})
            success = (
                response.get("success", "")
                if isinstance(response, dict)
                else ""
            )

            request_str = (
                json.dumps(
                    entry.get("clientRequest", {}),
                    ensure_ascii=False,
                    separators=(",", ":"),
                )
                if entry.get("clientRequest")
                else ""
            )
            response_str = (
                json.dumps(
                    response, ensure_ascii=False, separators=(",", ":")
                )
                if response
                else ""
            )

            ws.cell(row=row_idx, column=1, value=entry.get("id", row_idx - 1))
            ws.cell(row=row_idx, column=2, value=date_val)
            ws.cell(row=row_idx, column=3, value=route)
            ws.cell(row=row_idx, column=4, value=source)
            cell_success = ws.cell(
                row=row_idx, column=5, value=str(success)
            )
            if success is False:
                cell_success.font = Font(color="CC0000", bold=True)
            ws.cell(row=row_idx, column=6, value=request_str)
            ws.cell(row=row_idx, column=7, value=response_str)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 60
        ws.column_dimensions["G"].width = 50
        ws.freeze_panes = "A2"

        buf = BytesIO()
        wb.save(buf)
        file_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")

        filename = f"terminal_logs_{terminal.serial}_{date_str}.xlsx"

        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "datas": file_b64,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "res_model": terminal._name,
            "res_id": terminal.id,
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def _fetch_logs_for_date(self, terminal, date_str, exec_id):
        """Fetch log entries for a single date from the terminal API."""
        payload = {"data": {"date": date_str}}

        url = build_url(terminal.serial, "local", "logs")
        timestamp = int(datetime.utcnow().timestamp()) * 1000
        payload_clean = deep_clean_payload(payload)

        # Log the request
        log_payrillium_event(
            exec_id, "terminal_logs", "request",
            payload,
            success=True,
            env=self.env,
        )

        # Proven auth pattern: stringify → hash → wrap as string
        request_body = json.dumps(payload_clean, separators=(",", ":"))
        auth_hash = build_header_hash(self.env, payload_clean, timestamp)

        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Basic {auth_hash}",
            "timestamp": str(timestamp),
        }

        try:
            resp = requests.post(
                url, headers=headers,
                json={"data": request_body},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()

            entries = data.get("data", [])

            # Log the response
            log_payrillium_event(
                exec_id, "terminal_logs", "response",
                payload={
                    "date": date_str,
                    "serial": terminal.serial,
                    "success": data.get("success"),
                    "entries_count": len(entries),
                    "message": data.get("message", ""),
                },
                success=data.get("success", False),
                env=self.env,
            )

            if not data.get("success"):
                raise UserError(
                    f"Terminal returned error: "
                    f"{data.get('message', 'Unknown error')}"
                )

            return entries

        except requests.exceptions.RequestException as e:
            # Log the failure
            log_payrillium_event(
                exec_id, "terminal_logs", "response",
                payload={"date": date_str, "serial": terminal.serial},
                success=False,
                error_message=str(e),
                env=self.env,
            )
            raise
